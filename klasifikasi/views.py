from django.shortcuts import render, redirect
# from django.http import HttpResponseRedirect
from sklearn import svm #method untuk pross perhitungan klasifikasi
from sklearn.feature_extraction.text import TfidfTransformer, CountVectorizer #method untuk menghitung tfidf dan vsm
from sklearn import metrics #method untuk pembentukan matriks 1x1, 2x2, 3x3, ...
from sklearn.metrics import accuracy_score #method perhitungan akurasi
from sklearn.model_selection import KFold #Method perhitungan K-Fold
from sklearn.model_selection import train_test_split
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
import matplotlib.pyplot as plt
# from django.http import HttpResponse
import io
import urllib, base64
import pandas as pd
import nltk
import os
import nltk.corpus
import numpy as np #scientific computing untuk array N-dimenesi
import re #re = regular expression
import warnings
warnings.filterwarnings('ignore')

# import csv 

from django.shortcuts import render
import openpyxl

def index(request):
    if "GET" == request.method:
        return render(request, 'klasifikasi/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

            #CASE FOLDING + LOAD FILE DOKUMEN#
        dataSet = []#inisialisasi dataSet untuk menyimpan dokumen teks yang telah di looping
        for a in excel_data:#file dokumen yg akan di looping
            string = a[0]
            string = re.sub("[^a-zA-Z]", " ", string)#proses membuang karakter selain huruf diganti spasi/sub=substitusi(mereplace semua pola RE)     excel_data = excel_data.lower()#proses menjadikan kalimat huruf kecil
            string = string.lower()
            dataSet.append(string)#proses memasukan/mengupdate kalimat2 kedalam dataSet
        # print("Case Folding: \n", dataSet)

        # # #LOAD FILE STOPWORDS#
        stopword = []#inisialisasi dataSet untuk menyimpan dokumen teks
        s = open("klasifikasi/templates/id.stopwords.txt", "r+")#proses membuka file
        stop = s.read()#proses pembacaan file
        stop = re.sub("[^a-zA-Z]", " ", stop)#proses membuang karakter selain huruf diganti spasi/sub=substitusi(mereplace semua pola RE)
        stopword.append(stop)#proses memasukan/mengupdate kalimat2 kedalam stopword
        s.close()#proses menutup file
        # print("\nDaftar Stopword: \n", stopword)

        # TOKENIZING DOKUMEN#
        # range : berfungsi untuk mengembalikan deret integer berurut pada range yang ditentukan dari start sampai stop.

        bagOfWords = dataSet#insialisasi bank kata(bag of word) yang isinya sama dengan variabel dataSet
        for x in range(len(dataSet)):
        # for x in excel_data:#file dokumen yg akan di looping
            bagOfWords[x] = dataSet[x].split()#isi dari 'variabel dataSet' di pecah2 menjadi satuan kata lalu di copy ke sebuah variabel indeks ke-x
        print("\nTokenizing: \n", bagOfWords)

        # #TOKENIZING STOPWORDS#
        stopwords = stopword#insialisasi bank kata(variabel stopwords) yang isinya sama dengan variabel stopword
        for x in range(0,1):#file dokumen yg akan di looping
            stopwords[x] = stopword[x].split()#isi dari 'variabel word' di pecah2 menjadi satuan kata lalu di copy ke sebuah variabel indeks ke-x
        print("\nTokenizing Stopwords: \n", stopwords)

        #FILTERING#
        for x in range(len(dataSet)):#looping pada seluruh file dokumen abstrak
            for y in range(0, len(bagOfWords[x])):#looping pada setiap kata per dokumen
                for z in range(0, 780):#looping pada setiap kata stopwords
                    if(bagOfWords[x][y] == stopwords[0][z]):#proses membandingkan setiap kata per dokumen dgn setiap kata pada stopword
                        bagOfWords[x][y]=''#jika ditemukan kata yang tidak penting maka kata tsb dihapus
        print("\nFiltering: \n", bagOfWords)

        #KATA BERSIH/Mengembalikan kata2 yg sudah tidak ada kata yg 'tidak penting' menjadi kalimat utuh/dokumen#
        for i in range(0, len(bagOfWords)):#looping untuk seluruh kata pada bank kata
            bagOfWords[i] = filter(bool, bagOfWords[i])#menghapus kata yg kosong
            dataSet[i] = ' '.join(bagOfWords[i])#menggabungkan kata demi kata dengan sebuah pemisah spasi per dokumen
        print("\nKata Bersih: \n", dataSet)

        # factory = StemmerFactory()
        # stemmer = factory.create_stemmer()
        # for i in range(0, len(bagOfWords)):#looping untuk seluruh kata pada bank kata
        #     output = stemmer.stem(str(bagOfWords))
        # print("\nStemming : \n", output)

        #VSM & TFIDF#
        VSM = CountVectorizer().fit_transform(dataSet) #method vector space model dari library scikit learn melakukan perubahan menjadi sebuah vektor
        #tfidf = TfidfTransformer() #method tfidf dari library scikit learn di copy ke variabel tfidf
        TFIDF = TfidfTransformer().fit_transform(VSM) #method tfidf dari library scikit learn melakukan perubahan menjadi sebuah nilai
        #print (CountVectorizer().vocabulary)
        # print("\nVSM: \n", VSM)
        print("\n", VSM.todense())
        print("\nTFIDF: \n", TFIDF)
        #hhprint(TFIDF.todense())

# #KONVERSI LABEL# Data Latih
# #Kimia = 0, Fisika = 1, Biologi = 2
        # label_manual =  excel_data
        label = []
        for a in excel_data:
            label.append(a[1])
        label_manual = np.array(label)
#METHOD MENGHITUNG RATA2 AKURASI
        akurasi = []
        def avg_akurasi():
            # total = 0 ## pengosongan variabel
            # for i in range(10): ## looping 10x karena ada 10 fold
            #     total = total + akurasi[i]
            # rata2 = total / 10
            # avg_akurasi = (total / 10)
            print("-------------------------------------------------------------------------------------------------------") 
            print("Rata-rata akurasi keseluruhan adalah :", total / 10) ## cetak rata-rata akurasi

        data_prediksi = []
        data_uji = []
        data_latih = []
            
        kFoldCrossValidation = KFold(n_splits=10)#fungsi K-Fold Cross Validation melakukan insialisasi 10x iterasi
        for latih, uji in kFoldCrossValidation.split(TFIDF, label_manual):
            print("-----------------------------------------------------------------------")
            print("Banyak Data Latih: ", len(latih))
            print("Banyak Data Uji: ", len(uji))
            print("\nData Latih: \n", latih)
            print("\nData Uji: \n", uji)

            dataLatih1, dataUji1 = TFIDF[latih], TFIDF[uji]#proses inisialisasi dari masing2 data latih/uji dijadikan nilai tfidf lalu di copy ke variabel dataLatih/Uji1
            # label = []
            # for a in excel_data:
            #     label.append(a[1])
            # label_manual = np.array(label)

            dataLatih2, dataUji2 = label_manual[latih], label_manual[uji]#proses inisialisasi dari masing2 data latih/uji dibentuk ke label untuk proses prediksi lalu di copy ke variabel dataLatih/Uji2
            SVM = svm.SVC(kernel='linear').fit(dataLatih1, dataLatih2)#data latih melakukan proses pelatihan dengan algoritma SVM
            prediksi = SVM.predict(dataUji1)#proses prediksi dari data latih yang sudah tersimpan sebagai model

            print("\nHasil Prediksi: \n", prediksi)
            print("\nConfusion Matrix: \n", metrics.confusion_matrix(dataUji2, prediksi))#proses pembetukan metriks
            akurasi.append(accuracy_score(dataUji2, prediksi))
            print("\nAkurasi: ", accuracy_score(dataUji2, prediksi))
            print()
            label = ['Kimia', 'Fisika', 'Biologi']
            print(metrics.classification_report(dataUji2, prediksi, target_names=label))#proses pembentukan confusin matrix
            data_uji.append(uji)
            data_latih.append(latih)
            data_prediksi.append(prediksi)
            # metrics = metrics.classification_report(dataUji2, prediksi, target_names=label)

        # avg_akurasi()
        total = 0
        for i in range(10): ## looping 10x karena ada 10 fold
                total = total + akurasi[i]
        rata2 = total / 10


        np_prediksi = np.array(data_prediksi)
        data_hist = np_prediksi.ravel()
        plt.hist(data_hist)
        fig = plt.gcf()
        buf = io.BytesIO()
        fig.savefig(buf,format='png')
        buf.seek(0)
        string = base64.b64encode(buf.read())
        uri = urllib.parse.quote(string)
        # data_loop = zip(data_uji, data_prediksi)
        return render(request, 'klasifikasi/index.html', {"excel_data":excel_data, 'data' : uri, 'latih' : latih, 'uji' : uji, 'akurasi' : str(akurasi), 'rata2' : rata2, 'data_prediksi' : data_prediksi})
        # return HttpResponse(request, 'klasifikasi/index.html', list(akurasi))